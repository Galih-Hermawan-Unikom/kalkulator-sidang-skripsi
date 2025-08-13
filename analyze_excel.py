import json
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


def summarize_sheet(ws: Worksheet) -> Dict[str, Any]:
    dims = ws.calculate_dimension()
    max_row = ws.max_row
    max_col = ws.max_column

    # Column widths
    col_widths = {}
    for col_dim in ws.column_dimensions.values():
        if col_dim.index is None:
            continue
        col_widths[col_dim.index] = col_dim.width

    # Row heights
    row_heights = {}
    for row_idx, row_dim in ws.row_dimensions.items():
        row_heights[int(row_idx)] = row_dim.height

    # Merged cells
    merged: List[str] = [str(rng) for rng in ws.merged_cells.ranges]

    # Freeze panes
    freeze = None
    if ws.freeze_panes is not None:
        fp = ws.freeze_panes
        # In openpyxl, freeze_panes is typically a string like "A2"; handle both cases safely
        if isinstance(fp, str):
            freeze = fp
        else:
            freeze = getattr(fp, 'coordinate', str(fp))

    # Data validations
    validations = []
    if ws.data_validations:
        for dv in ws.data_validations.dataValidation:
            validations.append({
                'type': dv.type,
                'operator': dv.operator,
                'allowBlank': dv.allowBlank,
                'formula1': dv.formula1,
                'formula2': dv.formula2,
                'sqref': str(dv.sqref),
            })

    # Cells with formulas and constants
    formulas = []
    constants = []
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            if cell.value is None:
                continue
            coord = cell.coordinate
            if isinstance(cell.value, str) and cell.value.startswith('='):
                formulas.append({
                    'cell': coord,
                    'formula': cell.value,
                    'number_format': cell.number_format,
                })
            else:
                # keep short sample of constants for context (avoid dumping entire sheet)
                if len(constants) < 200:
                    constants.append({
                        'cell': coord,
                        'value': cell.value,
                        'number_format': cell.number_format,
                    })

    named_ranges = []
    # Named ranges are workbook-level; placeholder here to attach later if scoped.

    return {
        'title': ws.title,
        'dimension': dims,
        'max_row': max_row,
        'max_col': max_col,
        'column_widths': col_widths,
        'row_heights': row_heights,
        'merged_cells': merged,
        'freeze_panes': freeze,
        'validations': validations,
        'formulas': formulas,
        'constants_sample': constants,
    }


def analyze_workbook(path: Path) -> Dict[str, Any]:
    wb = load_workbook(path, data_only=False)

    sheets = [summarize_sheet(wb[s]) for s in wb.sheetnames]

    names = []
    if wb.defined_names is not None:
        try:
            iterable = None
            if hasattr(wb.defined_names, 'definedName'):
                iterable = wb.defined_names.definedName
            else:
                iterable = list(iter(wb.defined_names))
            for dn in iterable:
                try:
                    names.append({
                        'name': getattr(dn, 'name', str(dn)),
                        'attr_text': getattr(dn, 'attr_text', None),
                        'value': getattr(dn, 'value', None),
                        'comment': getattr(dn, 'comment', None),
                        'localSheetId': getattr(dn, 'localSheetId', None),
                    })
                except Exception:
                    names.append({'raw': str(dn)})
        except Exception:
            # Fallback: dump a string representation
            names = [{'raw': str(wb.defined_names)}]

    props = {
        'creator': getattr(wb.properties, 'creator', None),
        'lastModifiedBy': getattr(wb.properties, 'lastModifiedBy', None),
        'title': getattr(wb.properties, 'title', None),
        'description': getattr(wb.properties, 'description', None),
        'keywords': getattr(wb.properties, 'keywords', None),
        'category': getattr(wb.properties, 'category', None),
    }

    return {
        'file': str(path),
        'sheet_count': len(wb.sheetnames),
        'sheets': sheets,
        'defined_names': names,
        'properties': props,
    }


if __name__ == '__main__':
    xlsx = Path('perhitungan nilai sidang-2025-08-13.xlsx')
    if not xlsx.exists():
        raise SystemExit(f'File not found: {xlsx}')
    report = analyze_workbook(xlsx)
    print(json.dumps(report, indent=2, ensure_ascii=False))
